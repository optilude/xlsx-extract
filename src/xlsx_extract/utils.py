from typing import Tuple

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table
from openpyxl.cell import Cell
from openpyxl.utils.cell import quote_sheetname, range_to_tuple

from .range import Range

def get_range(ref : str, workbook : Workbook, worksheet : Worksheet = None) -> Range:
    """Get a Range by a reference, which can be defined name, a named table, or
    a cell/range reference.
    """

    defined_name = get_defined_name(workbook, worksheet, ref)
    named_table = None
    
    if defined_name is None:
        # Annoyingly, the named table API is at sheet level when it probably
        # should be at workbook level
        
        if worksheet is not None:
            named_table = get_named_table(worksheet, ref)
        else:
            for ws in workbook.worksheets:
                named_table = get_named_table(ws, ref)
                if named_table is not None:
                    worksheet = ws
                    break
    
    ref = \
        defined_name.attr_text if defined_name is not None \
        else named_table.ref if named_table is not None \
        else ref

    if worksheet is not None:
        ref = add_sheet_to_reference(worksheet, ref)
    
    # Failed to find the sheet
    if '!' not in ref:
        return None

    sheet_name, (c1, r1, c2, r2) = range_to_tuple(ref)

    # Malformed reference
    if None in (r1, c1, r2, c2,):
        return None
    
    # Might not be the same as `worksheet`
    sheet = workbook[sheet_name]

    cells = tuple(sheet.iter_rows(min_row=r1, min_col=c1, max_row=r2, max_col=c2))
    return Range(cells, defined_name=defined_name, named_table=named_table)

def get_defined_name(workbook : Workbook, worksheet : Worksheet, name : str) -> DefinedName:
    """Get a locally or globally defined name object
    """

    defined_name = None
    if worksheet is not None:
        defined_name = get_locally_defined_name(worksheet, name)

    if defined_name is None:
        defined_name = get_globally_defined_name(workbook, name)
    
    return defined_name

def get_locally_defined_name(worksheet : Worksheet, name : str) -> DefinedName:
    """Look up a defined name local to the worksheet
    """

    workbook = worksheet.parent
    
    sheet_id = workbook.index(worksheet)
    if name in workbook.defined_names.localnames(sheet_id):
        defined_name = workbook.defined_names.get(name, sheet_id)
        if defined_name is not None:
            return defined_name
    
    return None

def get_globally_defined_name(workbook : Workbook, name : str) -> DefinedName:
    """Look up a defined name global to the workbook
    """

    return workbook.defined_names.get(name, None)

def get_named_table(worksheet : Worksheet, name : str) -> Table:
    """Look up a named table
    """

    if worksheet is None or name not in worksheet.tables:
        return None

    return worksheet.tables[name]

def add_sheet_to_reference(worksheet : Worksheet, ref : str) -> str:
    """Add worksheet name to table if needed
    """
    if '!' not in ref:
        assert worksheet is not None, "Sheet must be given if reference does not contain a sheet name"
        ref = "%s!%s" % (quote_sheetname(worksheet.title), ref)
    return ref

def triangulate_cell(row : Cell, col : Cell) -> Cell:
    """Find the cell at the intersection of the row of `row`
    and the column of `col`.
    """
    assert row.parent is col.parent
    return row.parent.cell(row.row, col.column)

def copy_value(source : Cell, target : Cell):
    """Copy a single value from source to target
    """
    target.value = source.value

def resize_table(table : Range, rows : int, cols : int) -> Range:
    """Add or remove rows or columns at the end of of `table` so that it has the dimensions
    `rows` x `cols`. Return new table with the correct dimensions.
    """

    assert table is not None and not table.is_empty, \
        "Cannot resize an empty range"
    
    rows_delta = rows - table.rows
    cols_delta = cols - table.columns
    
    # No change
    if rows_delta == 0 and cols_delta == 0:
        return table

    # Add new rows to the bottom
    if rows_delta > 0:
        table.sheet.insert_rows(table.last_cell.row + 1, rows_delta)
    
    # Remove rows from the bottom
    if rows_delta < 0:
        table.sheet.delete_rows(table.first_cell.row + rows, -rows_delta)
    
    # Add new columns to the end
    if cols_delta > 0:
        table.sheet.insert_cols(table.last_cell.column + 1, cols_delta)
    
    # Remove columns from the top
    if cols_delta < 0:
        table.sheet.delete_cols(table.first_cell.column + cols, -cols_delta)
    
    new_range = tuple(
        table.sheet.iter_rows(
            min_row=table.first_cell.row,
            min_col=table.first_cell.column,
            max_row=table.first_cell.row + (rows - 1),
            max_col=table.first_cell.column + (cols - 1)
        )
    )
    
    new_table = Range(new_range, defined_name=table.defined_name, named_table=table.named_table)

    # Update defined name or named table reference if required

    if new_table.defined_name is not None:
        new_table.defined_name.attr_text = new_table.get_reference(absolute=True, use_sheet=True, use_defined_name=False, use_named_table=False)
    
    if table.named_table is not None:
        new_table.named_table.ref = new_table.get_reference(absolute=False, use_sheet=False, use_defined_name=False, use_named_table=False)
    
    return new_table

def update_table(source : Range, target : Range, expand : bool = True) -> Range:
    """Update target table with source table. Returns target (possibly expanded).
    """

    assert source is not None and not source.is_empty, \
        "Cannot copy an empty table (this is a bug - it should not happen)"
    
    assert target is not None and not target.is_empty, \
        "Cannot target an empty table (this is a bug - it should not happen)"

    if expand:
        target = resize_table(target, source.rows, source.columns)
    
    for source_row, target_row in zip(source.cells, target.cells):
        for source_cell, target_cell in zip(source_row, target_row):
            copy_value(source_cell, target_cell)
    
    return target

def extract_vector(table : Range, in_row : bool, index : int) -> Tuple[Cell]:
    """Get a tuple of the cells in the row at `index` if `in_row`, or in the
    column at `index` if not `in_row`.
    """
    return tuple(table.cells[index] if in_row else (r[index] for r in table.cells))

def align_vectors(
    source : Range, source_in_row : bool, source_idx : int,
    target : Range, target_in_row : bool, target_idx : int,
):
    """Replace a vector in `target` with a vector in `source` by matching
    labels in the first row/column of each.
    """

    assert source is not None and not source.is_empty, \
        "Cannot source from an empty table (this is a bug - it should not happen)"
    assert target is not None and not target.is_empty, \
        "Cannot target an empty table (this is a bug - it should not happen)"
    assert source_idx is not None, \
        "One of source row and column index must be set (this is a bug - it should not happen)"
    assert target_idx is not None, \
        "One of target row and column index must be set (this is a bug - it should not happen)"

    to_label = lambda s: s.strip().lower() if isinstance(s, (str, bytes,)) else s

    # Get the relevant source and target row or column into a single list
    source_vector = extract_vector(source, source_in_row, source_idx)
    target_vector = extract_vector(target, target_in_row, target_idx)

    assert source_idx >= 0 and source_idx < len(source_vector), \
        "Source row/column is outside of the target table (this is a bug - it should not happen)"
    assert target_idx >= 0 and target_idx < len(target_vector), \
        "Target row/column is outside of the target table (this is a bug - it should not happen)"

    # Find first row or column and use as labels
    source_labels = (to_label(c.value) for c in extract_vector(source, source_in_row, 0))
    target_labels = (to_label(c.value) for c in extract_vector(target, target_in_row, 0))

    source_lookup = dict(zip(source_labels, source_vector))

    # For each target label, find and copy the corresponding source cell
    for target_label, target_cell in zip(target_labels, target_vector):
        if target_label is not None:
            source_cell = source_lookup.get(target_label, None)
            if source_cell is not None:
                copy_value(source_cell, target_cell)


def replace_vector(
    source : Range, source_in_row : bool, source_idx : int,
    target : Range, target_in_row : bool, target_idx : int,
    expand : bool,
):
    """Replace a single row or column in target with a single row or column in source.
    """

    assert source is not None and not source.is_empty, \
        "Cannot source from an empty table (this is a bug - it should not happen)"
    assert target is not None and not target.is_empty, \
        "Cannot target an empty table (this is a bug - it should not happen)"
    assert source_idx is not None, \
        "One of source row and column index must be set (this is a bug - it should not happen)"
    assert target_idx is not None, \
        "One of target row and column index must be set (this is a bug - it should not happen)"

    # Get the relevant source and target row or column into a single list
    source_vector = extract_vector(source, source_in_row, source_idx)
    target_vector = extract_vector(target, target_in_row, target_idx)

    if expand:
        rows = len(source_vector) if not target_in_row else target.rows
        cols = len(source_vector) if target_in_row else target.columns

        target = resize_table(target, rows, cols)
        target_vector = extract_vector(target, target_in_row, target_idx)

    assert source_idx >= 0 and source_idx < len(source_vector), \
        "Source row/column is outside of the target table (this is a bug - it should not happen)"
    assert target_idx >= 0 and target_idx < len(target_vector), \
        "Target row/column is outside of the target table (this is a bug - it should not happen)"

    # Replace each value in the target bector with the corresponding value
    # in the target vector
    for source_cell, target_cell in zip(source_vector, target_vector):
        copy_value(source_cell, target_cell)

