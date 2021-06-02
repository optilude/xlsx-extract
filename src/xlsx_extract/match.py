import re

from enum import Enum
from typing import Any, Union, Tuple, Generator
from datetime import datetime, date, time
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

from .range import Range

from .utils import get_range

class Operator(Enum):

    EQUAL = "="
    NOT_EQUAL = "!="
    GREATER = ">"
    GREATER_EQUAL = ">="
    LESS = "<"
    LESS_EQUAL = "<="

    EMPTY = "is empty"
    NOT_EMPTY = "is not empty"

    REGEX = "regex"


@dataclass
class Comparator:
    """Parameters to find a single cell
    """

    operator : Operator
    value : Union[str, int, float, bool, date, time, datetime] = None

    def __post_init__(self):
        if self.operator == Operator.REGEX:
            assert type(self.value) is str, "Regular expression must be a string"

    def match(self, data : Union[str, int, float, bool, date, time, datetime]) -> Union[str, int, float, bool, date, time, datetime]:
        """Use the `operator` to compare `data` with `value`.

        Return value is `None` if not matched, or the matched item.
        For regex matches with match groups, the content of the first
        match group is returned (as a string).
        """

        value = self.value
        
        # note: datetime derives from date
        if isinstance(value, date) and not isinstance(value, datetime) and isinstance(data, datetime):
            value = datetime.fromordinal(value.toordinal())

        if self.operator == Operator.EMPTY:
            return "" if (
                (isinstance(data, str) and len(data) == 0) or
                (data is None)
            ) else None
        elif self.operator == Operator.NOT_EMPTY:
            return data if (
                (isinstance(data, str) and len(data) > 0) or
                (not isinstance(data, str) and data is not None)
            ) else None
        elif self.operator == Operator.EQUAL:
            try:
                return data if data == value else None
            except TypeError:
                return None
        elif self.operator == Operator.NOT_EQUAL:
            try:
                return data if data != value else None
            except TypeError:
                return None
        elif self.operator == Operator.GREATER:
            try:
                return data if data > value else None
            except TypeError:
                return None
        elif self.operator == Operator.GREATER_EQUAL:
            try:
                return data if data >= value else None
            except TypeError:
                return None
        elif self.operator == Operator.LESS:
            try:
                return data if data < value else None
            except TypeError:
                return None
        elif self.operator == Operator.LESS_EQUAL:
            try:
                return data if data <= value else None
            except TypeError:
                return None
        elif self.operator == Operator.REGEX:
            if not isinstance(data, (str, bytes)):
                return None
            
            match = re.search(value, data, re.IGNORECASE)
            if match is None:
                return None
            
            groups = match.groups()
            return groups[0] if len(groups) > 0 else data
        
        return None  # no match

@dataclass
class Match:

    name : str

    # What sheet are we on? Optional if `reference` is set
    sheet : Comparator = None

    # Search by cell/range reference (name or coordinate)
    reference : str = None

    def match(self, workbook : Workbook) -> Tuple[Range, Any]:
        """Match current parameters in worksheet and return a tuple of
        `(matched cells, matched value)`.
        """
        # Subclasses will implement
        raise NotImplementedError

    def get_sheet(self, workbook : Workbook) -> Tuple[Worksheet, str]:
        """Return the worksheet matching `self.sheet`, returning
        a tuple of the worksheet object and the matched title
        """
        for ws in workbook.worksheets:
            match = self.sheet.match(ws.title)
            if match is not None:
                return (ws, match)
        return (None, None)

    def find_by_reference(self, workbook : Workbook, worksheet : Worksheet = None) -> Range:
        """Find the cell or range matching `self.reference`. Always returns a
        tuple of tuples.
        """
        if self.reference is None or self.reference == "":
            return None

        return get_range(self.reference, workbook, worksheet)

@dataclass
class CellMatch(Match):
    """Target a single cell
    """
    
    # Search by cell contents
    value : Comparator = None
    
    # Find value in offset from the matched cell (can be positive or negative)
    row_offset : int = 0
    col_offset : int = 0

    # Define a search area
    min_row : int = None
    min_col : int = None
    max_row : int = None
    max_col : int = None

    def __post_init__(self):

        assert self.reference is not None or self.value is not None, \
            "%s: Either cell reference or cell value must be given to identify a cell" % self.name

        if self.reference is not None:
            assert self.value is None, "%s: Cell value cannot be specified if cell reference is given" % self.name
        
        if self.value is not None:
            # can be set in post-init by range match
            # assert self.sheet is not None, "%s: Sheet is required if matching by value" % self.name
            assert self.reference is None, "%s: Cell value cannot be specified if cell value is given" % self.name

    def match(self, workbook : Workbook) -> Tuple[Range, Any]:
        """Match a single cell
        """

        worksheet = None
        if self.sheet is not None:
            worksheet, _ = self.get_sheet(workbook)

        cell = None
        match = None

        if self.reference is not None:
            range = self.find_by_reference(workbook, worksheet)
            if range is not None and range.is_cell:
                cell = range.cell
        elif self.value is not None:
            cell, match = self.find_by_value(worksheet)
        
        if cell is None:
            return (None, None)

        if self.row_offset != 0 or self.col_offset != 0:
            cell = cell.offset(self.row_offset, self.col_offset)
        
        if match is None:
            match = cell.value

        return (
            Range(((cell,),)),
            match,
        )

    def find_by_value(self, worksheet : Worksheet) -> Tuple[Cell, Any]:
        """Search the worksheet for a cell by value comparator, returning
        a tuple of `(cell, match)` or `(None, None)`.
        """
        if self.value is None or worksheet is None:
            return (None, None)

        for row in self._iter_rows(worksheet):
            for cell in row:
                match_value = self.value.match(cell.value)
                if match_value is not None:
                    return (cell, match_value)
        
        return (None, None)
    
    def _iter_rows(self, worksheet : Worksheet) -> Generator[Tuple[Cell], None, None]:
        """Iterate over rows (list of cells) in the worksheet within the 
        min/max row/col boundaries.
        """
        return worksheet.iter_rows(
            min_row=self.min_row,
            max_row=self.max_row,
            min_col=self.min_col,
            max_col=self.max_col
        )

@dataclass
class RangeMatch(Match):

    # Find start of table by cell
    start_cell : CellMatch = None

    # Range extends until specified end cell
    end_cell : CellMatch = None
    
    # Range extends for a set number of rows and columns
    rows : int = None
    cols : int = None

    def __post_init__(self):

        assert self.reference is not None or self.start_cell is not None, \
            "%s: Either a reference or a start cell must be specified" % self.name

        if self.reference is not None:
            assert self.start_cell is None, "%s: Start cell cannot be specified if a reference is used" % self.name
            assert self.end_cell is None, "%s: End cell cannot be specified if a reference is used" % self.name
            assert self.rows is None, "%s: Row count cannot be specified if a reference is used" % self.name
            assert self.cols is None, "%s: Column count cannot be specified if a reference is used" % self.name
        
        if self.start_cell is not None:
            assert self.reference is None, "%s: A cell reference cannot be specified if a start cell is given" % self.name
            
            if self.sheet is not None:
                self.start_cell.sheet = self.sheet

            if self.end_cell is not None:
                assert self.rows is None and self.cols is None, "%s: Fixed row and column counts cannot be specified if an end cell is given" % self.name

                if self.sheet is not None:
                    self.end_cell.sheet = self.sheet

            if self.rows is not None:
                assert self.cols is not None, "%s: If a fixed row count is given, a fixed column count must also be specified" % self.name
            if self.cols is not None:
                assert self.rows is not None, "%s: If a fixed column count is given, a fixed row count must also be specified" % self.name

            if self.rows is not None and self.cols is not None:
                assert self.end_cell is None, "%s: An end cell cannot be specified if fixed row and column counts are given" % self.name

    def match(self, workbook : Workbook) -> Tuple[Range, Any]:
        """Match a range of cells
        """

        worksheet = None
        if self.sheet is not None:
            worksheet, _ = self.get_sheet(workbook)

        range = None
        match = None

        if self.reference is not None:
            range = self.find_by_reference(workbook, worksheet)
        elif self.start_cell is not None:
            if self.end_cell is not None:
                range, match = self.find_by_end_cell(workbook)
            elif self.rows is not None and self.cols is not None:
                range, match = self.find_by_dimensions(workbook)
            else:
                range, match = self.find_by_contiguous_region(workbook)
        
        return (range, match,)


    def find_by_end_cell(self, workbook : Workbook) -> Tuple[Range, Any]:
        """Find range by `self.start_cell` and `self.end_cell`.
        """
        if self.start_cell is None or self.end_cell is None:
            return (None, None,)

        start_cell_range, start_cell_match = self.start_cell.match(workbook)
        end_cell_range, _ = self.end_cell.match(workbook)
        
        if (
            (start_cell_range is None or not start_cell_range.is_cell) or
            (end_cell_range is None or not end_cell_range.is_cell) or
            (not start_cell_range.sheet is end_cell_range.sheet)
        ):
            return (None, None,)
        
        cells = start_cell_range.sheet["%s:%s" % (start_cell_range.cell.coordinate, end_cell_range.cell.coordinate,)]

        return (
            Range(cells),
            start_cell_match,
        )
    
    def find_by_dimensions(self, workbook : Workbook) -> Tuple[Range, Any]:
        """Find range by `self.start_cell`, `self.rows` and `self.cols`.
        """
        if self.start_cell is None or self.rows is None or self.cols is None:
            return (None, None,)

        start_cell_range, start_cell_match = self.start_cell.match(workbook)

        if start_cell_range is None or not start_cell_range.is_cell:
            return (None, None,)
        
        return (
            Range(
                tuple(start_cell_range.sheet.iter_rows(
                    min_row=start_cell_range.cell.row,
                    min_col=start_cell_range.cell.column,
                    max_row=start_cell_range.cell.row + (self.rows - 1),
                    max_col=start_cell_range.cell.column + (self.cols - 1)
                )),
            ),
            start_cell_match,
        )
    
    def find_by_contiguous_region(self, workbook : Workbook) -> Tuple[Range, Any]:
        """Find range contiguously from `self.start_cell`.
        """
        if self.start_cell is None:
            return (None, None,)

        start_cell_range, start_cell_match = self.start_cell.match(workbook)

        if start_cell_range is None or not start_cell_range.is_cell:
            return (None, None,)
        
        sheet = start_cell_range.sheet
        start_cell = start_cell_range.cell
        rows = 1
        cols = 1

        # find first blank column along first row
        # we use `iter_rows()` because `iter_cols()` isn't available in readonly mode!
        for r in sheet.iter_rows(min_row=start_cell.row, max_row=start_cell.row, min_col=start_cell.column + 1, values_only=True):
            for c in r:
                if c is None or c == "":
                    break
                cols += 1
        
        # find first blank row along first column
        for r in sheet.iter_rows(min_row=start_cell.row + 1, min_col=start_cell.column, max_col=start_cell.column, values_only=True):
            if len(r) == 0 or r[0] is None or r[0] == "":
                break
            rows += 1

        return (
            Range(
                tuple(sheet.iter_rows(
                    min_row=start_cell.row,
                    min_col=start_cell.column,
                    max_row=start_cell.row + (rows - 1),
                    max_col=start_cell.column + (cols - 1)
                ))
            ),
            start_cell_match
        )
