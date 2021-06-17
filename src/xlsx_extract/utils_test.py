import datetime
import os.path
import openpyxl

from .utils import (
    get_range,
    get_globally_defined_name,
    get_defined_name,
    get_named_table,
    add_sheet_to_reference,
    resize_table,
    triangulate_cell,
    copy_value,
    update_table,
    extract_vector,
    align_vectors,
    replace_vector,
)

from .range import Range

def get_test_workbook(filename='source.xlsx', data_only=True):
    filename = os.path.join(os.path.dirname(__file__), 'test_data', filename)
    return openpyxl.load_workbook(filename, data_only=data_only)

def test_get_range():
    wb = get_test_workbook()
    ws = wb['Report 2']

    assert get_range('A3', wb) is None
    assert get_range('A3', wb, ws).get_reference(absolute=False) == "'Report 2'!A3"
    
    assert get_range("'Report 1'!A3", wb).get_reference(absolute=False) == "'Report 1'!A3"
    assert get_range("'Report 1'!A3", wb, ws).get_reference(absolute=False) == "'Report 1'!A3"

    assert get_range("'Report 1'!A3:B4", wb).get_reference(absolute=False) == "'Report 1'!A3:B4"
    assert get_range("'Report 1'!A3:B4", wb, ws).get_reference(absolute=False) == "'Report 1'!A3:B4"

    assert get_range("'Report 1'!A3", wb).get_reference(absolute=False) == "'Report 1'!A3"
    assert get_range("'Report 1'!A3", wb, ws).get_reference(absolute=False) == "'Report 1'!A3"

    assert get_range("RangleTable", wb).get_reference() == "RangleTable"
    assert get_range("RangleTable", wb, ws).get_reference() == "RangleTable"
    assert get_range("RangleTable", wb, ws).get_reference(absolute=False, use_named_table=False) == "'Report 2'!B10:E13"

    assert get_range("PROFIT_RANGE", wb).get_reference() == "PROFIT_RANGE"
    assert get_range("PROFIT_RANGE", wb, ws).get_reference(absolute=False) == "PROFIT_RANGE"
    assert get_range("PROFIT_RANGE", wb, ws).get_reference(absolute=False, use_defined_name=False) == "'Report 3'!A1:E5"

    assert get_range('NotFound', wb) is None
    assert get_range('NotFound', wb, ws) is None

def test_get_globally_defined_name():
    wb = get_test_workbook()

    assert get_globally_defined_name(wb, "PROFIT_RANGE") is not None
    assert get_globally_defined_name(wb, "FOOBAR") is None

def test_get_defined_name():
    wb = get_test_workbook()
    ws = wb['Report 1']

    assert get_defined_name(wb, ws, "PROFIT_RANGE") is not None
    assert get_defined_name(wb, ws, "FOOBAR") is None

def test_get_named_table():
    wb = get_test_workbook()
    ws = wb['Report 2']

    assert get_named_table(ws, "RangleTable") is not None
    assert get_named_table(ws, "NotFound") is None

def test_add_sheet_to_reference():
    wb = get_test_workbook()
    ws = wb['Report 1']

    assert add_sheet_to_reference(ws, "B3:C4") == "'Report 1'!B3:C4"

def test_triangulate_cell():
    wb = get_test_workbook()
    ws = wb['Report 1']

    row = ws.cell(row=3, column=5)
    col = ws.cell(row=6, column=8)

    cell = triangulate_cell(row, col)

    assert cell.row == 3
    assert cell.column == 8

def test_copy_value():
    wb = get_test_workbook()
    ws = wb['Report 1']

    c1 = ws.cell(row=3, column=2)
    c2 = ws.cell(row=4, column=6)

    assert c1.value == "Date"
    assert c2.value != "Date"

    copy_value(c1, c2)

    c1 = ws.cell(row=3, column=2)
    c2 = ws.cell(row=4, column=6)
    
    assert c1.value == "Date"
    assert c2.value == "Date"

class TestResizeTable:

    def get_table(self):
        wb = get_test_workbook()
        ws = wb['Report 1']

        # Full table: 5 rows by 4  columns
        assert Range(ws['B5:F9']).get_values() == (
            (None, 'Jan', 'Feb', 'Mar', 'Apr',),
            ('Alpha', 1.5, 6, 11, 4.6,),
            ('Beta', 2, 7, 12, 4.7,),
            ('Delta', 2.5, 8, 13, 4.8,),
            ('Gamma', 3, 9, 14, 4.9,),
        )

        # Small table: 3 rows by 2 columns
        table = Range(ws['C5:D7'])
        assert table.get_values() == (
            ('Jan', 'Feb',),
            (1.5, 6,),
            (2, 7,),
        )
        
        return ws, table

    def test_add_rows(self):
        ws, table = self.get_table()

        # Resize small table: add two rows
        new_table = resize_table(table, rows=5, cols=2)
        
        assert new_table.get_values() == (
            ('Jan', 'Feb',),
            (1.5, 6,),
            (2, 7,),
            (None, None,),
            (None, None,),
        )

        assert Range(ws['B5:F11']).get_values() == (
            (None, 'Jan', 'Feb', 'Mar', 'Apr',),
            ('Alpha', 1.5, 6, 11, 4.6,),
            ('Beta', 2, 7, 12, 4.7,),
            (None, None, None, None, None,),
            (None, None, None, None, None,),
            ('Delta', 2.5, 8, 13, 4.8,),
            ('Gamma', 3, 9, 14, 4.9,),
        )
    
    def test_remove_rows(self):
        ws, table = self.get_table()

        # Resize small table: remove one row
        new_table = resize_table(table, rows=2, cols=2)

        assert new_table.get_values() == (
            ('Jan', 'Feb',),
            (1.5, 6,),
        )

        assert Range(ws['B5:F8']).get_values() == (
            (None, 'Jan', 'Feb', 'Mar', 'Apr',),
            ('Alpha', 1.5, 6, 11, 4.6,),
            ('Delta', 2.5, 8, 13, 4.8,),
            ('Gamma', 3, 9, 14, 4.9,),
        )
    
    def test_add_cols(self):
        ws, table = self.get_table()

        # Resize small table: add two columns
        new_table = resize_table(table, rows=3, cols=4)

        assert new_table.get_values() == (
            ('Jan', 'Feb', None, None,),
            (1.5, 6, None, None,),
            (2, 7, None, None,),
        )

        assert Range(ws['B5:H9']).get_values() == (
            (None, 'Jan', 'Feb', None, None, 'Mar', 'Apr',),
            ('Alpha', 1.5, 6, None, None, 11, 4.6,),
            ('Beta', 2, 7, None, None, 12, 4.7,),
            ('Delta', 2.5, 8, None, None, 13, 4.8,),
            ('Gamma', 3, 9, None, None, 14, 4.9,),
        )
    
    def test_remove_cols(self):
        ws, table = self.get_table()

        # Resize small table: remove one column
        new_table = resize_table(table, rows=3, cols=1)

        assert new_table.get_values() == (
            ('Jan',),
            (1.5,),
            (2,),
        )

        assert Range(ws['B5:E9']).get_values() == (
            (None, 'Jan', 'Mar', 'Apr',),
            ('Alpha', 1.5, 11, 4.6,),
            ('Beta', 2, 12, 4.7,),
            ('Delta', 2.5, 13, 4.8,),
            ('Gamma', 3, 14, 4.9,),
        )
    
    def test_change_both_dimensions(self):
        ws, table = self.get_table()
        
        # Remove one row, add two columns
        new_table = resize_table(table, rows=2, cols=4)

        assert new_table.get_values() == (
            ('Jan', 'Feb', None, None,),
            (1.5, 6, None, None,),
        )

        assert Range(ws['B5:H8']).get_values() == (
            (None, 'Jan', 'Feb', None, None, 'Mar', 'Apr',),
            ('Alpha', 1.5, 6, None, None, 11, 4.6,),
            ('Delta', 2.5, 8, None, None, 13, 4.8,),
            ('Gamma', 3, 9, None, None, 14, 4.9,),
        )
    
    def test_resize_defined_name_table(self):
        wb = get_test_workbook()

        table = get_range('PROFIT_RANGE', wb)

        assert table.get_values() == (
            (None, 'Profit', None, 'Loss', None,),
            (None, '£',	'Plan', '£', 'Plan',),
            ('Alpha', 100, 100, 50, 20,),
            ('Beta', 200, 150, 50, 20,),
            ('Delta', 300, 350, 50, 20,),
        )

        # add two columns, remove one row
        new_table = resize_table(table, rows=4, cols=7)

        assert new_table.get_values() == (
            (None, 'Profit', None, 'Loss', None, None, None,),
            (None, '£',	'Plan', '£', 'Plan', None, None,),
            ('Alpha', 100, 100, 50, 20, None, None,),
            ('Beta', 200, 150, 50, 20, None, None,),
        )

        # check that the named range now resolves to the new table
        confirm_table = get_range('PROFIT_RANGE', wb)

        assert confirm_table.get_values() == (
            (None, 'Profit', None, 'Loss', None, None, None,),
            (None, '£',	'Plan', '£', 'Plan', None, None,),
            ('Alpha', 100, 100, 50, 20, None, None,),
            ('Beta', 200, 150, 50, 20, None, None,),
        )
    
    def test_resize_named_table(self):
        wb = get_test_workbook()
        ws = wb['Report 2']
        
        table = get_range('RangleTable', wb, ws)

        assert table.get_values() == (
            ('Name', 'Date', 'Range', 'Price',),
            ('Bill', datetime.datetime(2021, 1, 1), 9, 15,),
            ('Bob', datetime.datetime(2021, 3, 2), 14, 18,),
            ('Joan', datetime.datetime(2021, 6, 5), 13, 99,),
        )

        # Remove a column, add two rows

        new_table = resize_table(table, rows=6, cols=3)

        assert new_table.get_values() == (
            ('Name', 'Date', 'Range',),
            ('Bill', datetime.datetime(2021, 1, 1), 9,),
            ('Bob', datetime.datetime(2021, 3, 2), 14,),
            ('Joan', datetime.datetime(2021, 6, 5), 13,),
            (None, None, None,),
            (None, None, None,),
        )

        # Check the named reference was updated

        confirm_table = get_range('RangleTable', wb, ws)

        assert confirm_table.get_values() == (
            ('Name', 'Date', 'Range',),
            ('Bill', datetime.datetime(2021, 1, 1), 9,),
            ('Bob', datetime.datetime(2021, 3, 2), 14,),
            ('Joan', datetime.datetime(2021, 6, 5), 13,),
            (None, None, None,),
            (None, None, None,),
        )

def test_update_table_without_expanding():
    source_wb = get_test_workbook('source.xlsx')
    target_wb = get_test_workbook('target.xlsx', data_only=False)

    source = Range(source_wb['Report 1']['B5:F9'])
    target = Range(target_wb['Summary']['B7:E9'])

    assert source.get_values() == (
        (None, 'Jan', 'Feb', 'Mar', 'Apr',),
        ('Alpha', 1.5, 6, 11, 4.6,),
        ('Beta', 2, 7, 12, 4.7,),
        ('Delta', 2.5, 8, 13, 4.8,),
        ('Gamma', 3, 9, 14, 4.9,),
    )

    assert target.get_values() == (
        (None, 'Alpha', 'Delta', 'Beta',),
        ('Profit', None, None, None,),
        ('Loss', None, None, None,),
    )

    assert target_wb['Summary']['B11'].value == "Area"

    new_target = update_table(source, target, False)

    assert new_target.get_values() == (
        (None, 'Jan', 'Feb', 'Mar',),
        ('Alpha', 1.5, 6, 11,),
        ('Beta', 2, 7, 12,),
    )

    confirm_target = Range(target_wb['Summary']['B7:E9'])

    assert confirm_target.get_values() == (
        (None, 'Jan', 'Feb', 'Mar',),
        ('Alpha', 1.5, 6, 11,),
        ('Beta', 2, 7, 12,),
    )

    assert target_wb['Summary']['B11'].value == "Area"

def test_update_table_and_expand():
    source_wb = get_test_workbook('source.xlsx')
    target_wb = get_test_workbook('target.xlsx', data_only=False)

    source = Range(source_wb['Report 1']['B5:F9'])
    target = Range(target_wb['Summary']['B7:E9'])

    assert source.get_values() == (
        (None, 'Jan', 'Feb', 'Mar', 'Apr',),
        ('Alpha', 1.5, 6, 11, 4.6,),
        ('Beta', 2, 7, 12, 4.7,),
        ('Delta', 2.5, 8, 13, 4.8,),
        ('Gamma', 3, 9, 14, 4.9,),
    )

    assert target.get_values() == (
        (None, 'Alpha', 'Delta', 'Beta',),
        ('Profit', None, None, None,),
        ('Loss', None, None, None,),
    )

    # Will be pushed down
    assert target_wb['Summary']['B11'].value == "Area"

    new_target = update_table(source, target, True)

    assert new_target.get_values() == (
        (None, 'Jan', 'Feb', 'Mar', 'Apr',),
        ('Alpha', 1.5, 6, 11, 4.6,),
        ('Beta', 2, 7, 12, 4.7,),
        ('Delta', 2.5, 8, 13, 4.8,),
        ('Gamma', 3, 9, 14, 4.9,),
    )

    confirm_target = Range(target_wb['Summary']['B7:F11'])

    assert confirm_target.get_values() == (
        (None, 'Jan', 'Feb', 'Mar', 'Apr',),
        ('Alpha', 1.5, 6, 11, 4.6,),
        ('Beta', 2, 7, 12, 4.7,),
        ('Delta', 2.5, 8, 13, 4.8,),
        ('Gamma', 3, 9, 14, 4.9,),
    )

    # Has been pushed down
    assert target_wb['Summary']['B13'].value == "Area"

def test_extract_vector():
    wb = get_test_workbook()

    table = Range(wb['Report 1']['B5:F9'])

    assert table.get_values() == (
        (None, 'Jan', 'Feb', 'Mar', 'Apr',),
        ('Alpha', 1.5, 6, 11, 4.6,),
        ('Beta', 2, 7, 12, 4.7,),
        ('Delta', 2.5, 8, 13, 4.8,),
        ('Gamma', 3, 9, 14, 4.9,),
    )
    
    assert tuple(c.value for c in extract_vector(table, in_row=True, index=1)) == ('Alpha', 1.5, 6, 11, 4.6,)
    assert tuple(c.value for c in extract_vector(table, in_row=False, index=1)) == ('Jan', 1.5, 2, 2.5, 3,)

def test_align_vectors():
    source_wb = get_test_workbook('source.xlsx')
    target_wb = get_test_workbook('target.xlsx', data_only=False)

    source = Range(source_wb['Report 1']['B5:F9'])
    target = Range(target_wb['Summary']['B7:E9'])

    assert source.get_values() == (
        (None,   'Jan', 'Feb', 'Mar', 'Apr',),
        ('Alpha', 1.5,     6,     11,   4.6,),
        ('Beta',    2,     7,     12,   4.7,),
        ('Delta', 2.5,     8,     13,   4.8,),
        ('Gamma',   3,     9,     14,   4.9,),
    )

    assert target.get_values() == (
        (None,     'Alpha', 'Delta', 'Beta',),
        ('Profit',    None,    None,  None,),
        ('Loss',      None,    None,  None,),
    )

    # Align third column of source to second row of target
    align_vectors(source, False, 2, target, True, 1)

    assert Range(target_wb['Summary']['B7:E9']).get_values() == (
        (None,     'Alpha', 'Delta', 'Beta',),
        ('Profit',       6,       8,      7,),
        ('Loss',      None,    None,   None,),
    )

def test_replace_vector():
    source_wb = get_test_workbook('source.xlsx')
    target_wb = get_test_workbook('target.xlsx', data_only=False)

    source = Range(source_wb['Report 1']['B5:F9'])
    target = Range(target_wb['Summary']['B7:E9'])

    assert source.get_values() == (
        (None,   'Jan', 'Feb', 'Mar', 'Apr',),
        ('Alpha', 1.5,     6,     11,   4.6,),
        ('Beta',    2,     7,     12,   4.7,),
        ('Delta', 2.5,     8,     13,   4.8,),
        ('Gamma',   3,     9,     14,   4.9,),
    )

    assert target.get_values() == (
        (None,     'Alpha', 'Delta', 'Beta',),
        ('Profit',    None,    None,  None,),
        ('Loss',      None,    None,  None,),
    )

    # Replace third column of source to second row of target
    replace_vector(source, False, 2, target, True, 1, expand=False)

    assert Range(target_wb['Summary']['B7:F9']).get_values() == (
        (None,   'Alpha', 'Delta', 'Beta', None,),
        ('Feb',   6,            7,      8, None,),
        ('Loss',  None,      None,   None, None,),
    )

def test_replace_vector_expand():
    source_wb = get_test_workbook('source.xlsx')
    target_wb = get_test_workbook('target.xlsx', data_only=False)

    source = Range(source_wb['Report 1']['B5:F9'])
    target = Range(target_wb['Summary']['B7:E9'])

    assert source.get_values() == (
        (None,   'Jan', 'Feb', 'Mar', 'Apr',),
        ('Alpha', 1.5,     6,     11,   4.6,),
        ('Beta',    2,     7,     12,   4.7,),
        ('Delta', 2.5,     8,     13,   4.8,),
        ('Gamma',   3,     9,     14,   4.9,),
    )

    assert target.get_values() == (
        (None,     'Alpha', 'Delta', 'Beta',),
        ('Profit',    None,    None,  None,),
        ('Loss',      None,    None,  None,),
    )

    # Replace third column of source to second row of target
    replace_vector(source, False, 2, target, True, 1, expand=True)

    assert Range(target_wb['Summary']['B7:F9']).get_values() == (
        (None,   'Alpha', 'Delta', 'Beta', None,),
        ('Feb',   6,            7,      8,    9,),
        ('Loss',  None,      None,   None, None,),
    )
