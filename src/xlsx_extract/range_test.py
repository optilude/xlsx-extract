import datetime
import os.path
import openpyxl

from . import range, utils

def get_test_workbook():
    filename = os.path.join(os.path.dirname(__file__), 'test_data', 'source.xlsx')
    return openpyxl.load_workbook(filename, data_only=True)

class TestRange:

    def test_empty(self):
        r = range.Range(())
        
        assert r.is_empty
        assert not r.is_cell
        assert not r.is_range

        assert r.workbook is None
        assert r.sheet is None
        assert r.cell is None
        assert r.first_cell is None
        assert r.last_cell is None
        assert r.rows == 0
        assert r.columns == 0

        assert r.get_reference() is None
        assert r.get_values() == ()
    
    def test_single_cell(self):
        wb = get_test_workbook()
        ws = wb['Report 1']
        cells = ws['B3:B3']

        r = range.Range(cells)

        assert not r.is_empty
        assert r.is_cell
        assert not r.is_range

        assert r.workbook is wb
        assert r.sheet is ws
        assert r.cell is cells[0][0]
        assert r.first_cell is cells[0][0]
        assert r.last_cell is cells[0][0]
        assert r.rows == 1
        assert r.columns == 1

        assert r.get_reference() == "'Report 1'!$B$3"
        assert r.get_reference(absolute=False) == "'Report 1'!B3"
        assert r.get_reference(use_sheet=False) == "$B$3"

        assert r.get_values() == (('Date',),)
    
    def test_range_cell(self):
        wb = get_test_workbook()
        ws = wb['Report 1']
        cells = ws['B2:C3']

        r = range.Range(cells)

        assert not r.is_empty
        assert not r.is_cell
        assert r.is_range

        assert r.workbook is wb
        assert r.sheet is ws
        assert r.cell is None
        assert r.first_cell is cells[0][0]
        assert r.last_cell is cells[-1][-1]
        assert r.rows == 2
        assert r.columns == 2

        assert r.get_reference() == "'Report 1'!$B$2:$C$3"
        assert r.get_reference(absolute=False) == "'Report 1'!B2:C3"
        assert r.get_reference(use_sheet=False) == "$B$2:$C$3"
        
        assert r.get_values() == (
            (None, None,),
            ('Date', datetime.datetime(2021, 5, 1),),
        )
    
    def test_defined_name(self):
        wb = get_test_workbook()

        r = utils.get_range("PROFIT_RANGE", wb)

        assert not r.is_empty
        assert not r.is_cell
        assert r.is_range

        assert r.workbook is wb
        assert r.cell is None
        assert r.rows == 5
        assert r.columns == 5

        assert r.get_reference() == "PROFIT_RANGE"
        assert r.get_reference(use_defined_name=False) == "'Report 3'!$A$1:$E$5"
        assert r.get_reference(use_defined_name=False, absolute=False) == "'Report 3'!A1:E5"
        assert r.get_reference(use_defined_name=False, use_sheet=False) == "$A$1:$E$5"

        assert r.get_values() == (
            (None, 'Profit', None, 'Loss', None,),
            (None, '£',	'Plan', '£', 'Plan',),
            ('Alpha', 100, 100, 50, 20,),
            ('Beta', 200, 150, 50, 20,),
            ('Delta', 300, 350, 50, 20,),
        )
        
    
    def test_named_table(self):
        wb = get_test_workbook()
        ws = wb['Report 2']
        
        r = utils.get_range('RangleTable', wb, ws)

        assert not r.is_empty
        assert not r.is_cell
        assert r.is_range

        assert r.workbook is wb
        assert r.sheet is ws
        assert r.cell is None
        assert r.rows == 4
        assert r.columns == 4

        assert r.get_reference() == "RangleTable"
        assert r.get_reference(use_named_table=False) == "'Report 2'!$B$10:$E$13"
        assert r.get_reference(use_named_table=False, absolute=False) == "'Report 2'!B10:E13"
        assert r.get_reference(use_named_table=False, use_sheet=False) == "$B$10:$E$13"
    
        assert r.get_values() == (
            ('Name', 'Date', 'Range', 'Price',),
            ('Bill', datetime.datetime(2021, 1, 1), 9, 15,),
            ('Bob', datetime.datetime(2021, 3, 2), 14, 18,),
            ('Joan', datetime.datetime(2021, 6, 5), 13, 99,),
        )