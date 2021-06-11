import os
import openpyxl

from dataclasses import dataclass
from string import Template
from typing import Any, Dict, List, Match, Tuple, Callable

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException

from zipfile import BadZipFile

from .range import Range
from .match import Comparator, RangeMatch, CellMatch, Operator
from .target import Target

class GlobalKeys:
    """Keys used outside a target/match block
    """

    # Optional: Sets search directory for source files
    DIRECTORY = "directory"

    # Set source file
    FILE = "file"

class MatchKeys:
    """Keys used for any type of match
    """

    # Starts a new match block
    NAME = "name"

    # Sheet to search
    SHEET = "sheet"

class CellMatchKeys(MatchKeys):
    """Keys specific to cell matches
    """

    # A reference to a single cell (coordinate or name)
    CELL_REFERENCE = "cell"

    VALUE = "value"

    MIN_ROW = "min row"
    MAX_ROW = "max row"
    MIN_COL = "min column"
    MAX_COL = "max column"

    ROW_OFFSET = "row offset"
    COL_OFFSET = "column offset"

class RangeMatchKeys(MatchKeys):
    """Keys specific to range matches
    """

    # A reference to a range, defined name pointing to a table, or named table
    TABLE_REFERENCE = "table"

    # Size of matched table
    ROWS = "rows"
    COLS = "columns"

class TargetKeys:

    # Target reference
    TARGET_CELL = "target cell"
    TARGET_TABLE = "target table"

    # Whether to align source and target table
    ALIGN = "align"

    # Whether to resize target table to match source (or truncate)
    EXPAND = "expand"

class Prefix:
    """Prefixes used for finding cells within a table
    """

    START = "start"
    END = "end"
    
    SOURCE_ROW = "source row"
    SOURCE_COL = "source column"
    
    TARGET_ROW = "target row"
    TARGET_COL = "target column"

    @classmethod
    def none(cls, s):
        return s

    @classmethod
    def start(cls, s):
        return "%s %s" % (cls.START, s)
    
    @classmethod
    def end(cls, s):
        return "%s %s" % (cls.END, s)
    
    @classmethod
    def source_row(cls, s):
        return "%s %s" % (cls.SOURCE_ROW, s)
    
    @classmethod
    def source_col(cls, s):
        return "%s %s" % (cls.SOURCE_COL, s)

    @classmethod
    def target_row(cls, s):
        return "%s %s" % (cls.TARGET_ROW, s)
    
    @classmethod
    def target_col(cls, s):
        return "%s %s" % (cls.TARGET_COL, s)
    
Operators = {
    "is": Operator.EQUAL,
    "=": Operator.EQUAL,
    "==": Operator.EQUAL,
    "is not": Operator.NOT_EQUAL,
    "!=": Operator.NOT_EQUAL,
    "matches": Operator.REGEX,
    "regex": Operator.REGEX,
    "<": Operator.LESS,
    "<=": Operator.LESS_EQUAL,
    ">": Operator.GREATER,
    ">=": Operator.GREATER_EQUAL,
    "is empty": Operator.EMPTY,
    "empty": Operator.EMPTY,
    "is not empty": Operator.NOT_EMPTY,
    "not empty": Operator.NOT_EMPTY,
}

class LowerDict(dict):
    """Dict where keys are always fetched in lowercase
    """

    def __getitem__(self, name):
        return super().__getitem__(name.lower())

class VariableTemplate(Template):
    """Case-insensitive version of String.Template
    """

    def safe_substitute(self, mapping=None, **kws):
        if mapping is None:
            mapping = {}
        m = LowerDict((k.lower(), v) for k, v in mapping.items())
        m.update(LowerDict((k.lower(), v) for k, v in kws.items()))
        return super().safe_substitute(m)

@dataclass
class Action:
    """Used for logging actions
    """

    name : str
    success : bool
    message : str

    comparator : Comparator = None
    match : Match = None
    target : Target = None

    def __str__(self):
        return "%s: %s - %s" % (
            self.name.capitalize(),
            "Success" if self.success else "Failed",
            self.message
        )


def run(target_workbook : Workbook, source_directory : str, source_file : str = None, config_sheet : str = "Config") -> List[Action]:
    """Load configuration from the given sheet in the target workbook and execute
    each step. Returns a history of what's happened. Tries pretty hard not to raise
    exceptions.
    """

    history = []
    variables = {}
    source_workbook = None

    if config_sheet not in target_workbook.sheetnames:
        history.append(Action("Configuration", False, "Configuration sheet `%s` does not exist in target workbook" % config_sheet))
        return history

    if source_file is not None:
        try:
            source_workbook = load_workbook(source_file, data_only=True)
        except (InvalidFileException, FileNotFoundError, BadZipFile,) as e:
            history.append(Action(GlobalKeys.FILE, False, str(e)))
        else:
            history.append(Action(GlobalKeys.FILE, True, "Opened %s" % source_file))
            variables[GlobalKeys.FILE] = source_file

    # Find contiguous range matches for each of "directory", "file" or "name",
    # until sheet is exhausted, and put into the `blocks` list

    block_match = RangeMatch(
        name="block",
        sheet=Comparator(Operator.EQUAL, config_sheet),
        start_cell=CellMatch("key",
            value=Comparator(
                Operator.REGEX, r'^\s*(' + '|'.join(
                    (GlobalKeys.DIRECTORY, GlobalKeys.FILE, MatchKeys.NAME,)
                ) + r')\s*$'
            ),
            min_row=1,
        )
    )

    history.append(Action("Extract", True, "Starting extract"))
    
    num_blocks = 0
    while (match := block_match.match(target_workbook)) != (None, None,):
        block_range, _ = match

        # prepare for next match (put this up top so we can use `continue` safely!)
        block_match.start_cell.min_row = block_range.last_cell.row + 1

        block = None
        source_match = None
        target = None

        try:
            block = parse_block(block_range, variables)
        except AssertionError as e:
            # Block contained an explicit parsing error (e.g. invalid operator)
            history.append(Action(block_range.first_cell.value, False, str(e)))
            continue
        
        if block is None:
            # Block was not a block after all - ignore
            continue
        
        if GlobalKeys.DIRECTORY in block:
            try:
                source_directory = extract_directory(block)
            except AssertionError as e:
                # Malformed block
                history.append(Action(GlobalKeys.DIRECTORY, False, str(e)))
                continue

            if source_directory is None:
                # Block was not a block after all - ignore
                continue
            
            history.append(Action(GlobalKeys.DIRECTORY, True, "Found %s" % source_directory, comparator=block[GlobalKeys.DIRECTORY]))
            variables[GlobalKeys.DIRECTORY] = source_directory

        if GlobalKeys.FILE in block:
            source_file = None
            file_match = None

            try:
                source_file, file_match = extract_filename(block, source_directory)
            except AssertionError as e:
                # Directory or file not found, or malformed block
                history.append(Action(GlobalKeys.FILE, False, str(e)))
                continue

            if source_file is None:
                # Block was not a block after all - ignore
                continue
            
            try:
                source_workbook = openpyxl.load_workbook(source_file, data_only=True)
            except (InvalidFileException, FileNotFoundError,) as e:
                history.append(Action(GlobalKeys.FILE, False, str(e), comparator=block[GlobalKeys.FILE]))
                continue
            
            history.append(Action(GlobalKeys.FILE, True, "Obtained %s" % source_file, comparator=block[GlobalKeys.FILE]))
            variables[GlobalKeys.FILE] = file_match

        if MatchKeys.NAME in block:
            block_name = block[MatchKeys.NAME].value

            # Don't parse target blocks if we don't yet have a file
            if source_workbook is None:
                history.append(Action(block_name, False, "No source workbook set ahead of %s" % block_name))
                continue

            try:
                source_match = extract_source_match(block)
            except AssertionError as e:
                # An assertion failed during match construction
                history.append(Action(block_name, False, str(e)))
                continue

            if source_match is None:
                history.append(Action(block_name, False, "Could not extract source match from block %s" % block_name))
                continue
        
            try:
                target = extract_target(block, source_match)
            except AssertionError as e:
                # An assertion failed during target construction
                history.append(Action(block_name, False, str(e)))
                continue
            
            match_range = None
            match_value = None

            # Source match only (used to define variables)
            if target is None:
                try:
                    match_range, match_value = source_match.match(source_workbook)
                except AssertionError as e:
                    # An assertion failed during match execution
                    history.append(Action(block_name, False, str(e), match=source_match))
                    continue
            # Target with source match
            else:
                try:
                    match_range, match_value = target.extract(source_workbook, target_workbook)
                except AssertionError as e:
                    # An assertion failed during target execution
                    history.append(Action(block_name, False, str(e), match=source_match, target=target))
                    continue
            
            if match_range is None:
                history.append(Action(block_name, False, "Failed to match", match=source_match, target=target))
            else:
                history.append(Action(block_name, True, "Matched", match=source_match, target=target))
            
            num_blocks += 1
            if match_value is not None:
                variables[block_name] = match_value
    
    history.append(Action("Extract", all(a.success for a in history), "Extracted %d blocks" % num_blocks))
    return history

def parse_block(match_range : Range, variables : Dict[str, Any]) -> Dict[str, Comparator]:
    """Turn a 3-column range into a dict of lowercase string keys to comparators
    """

    if match_range.is_empty or not match_range.is_range or match_range.columns < 3:
        return None

    block = {}

    for row in match_range.get_values():

        if len(row) < 3:
            continue

        name, operator, value = row[:3]
        if (
            not isinstance(name, (str, bytes,)) or name == "" or
            not isinstance(operator, (str, bytes,)) or operator == ""
        ):
            continue
        
        value = interpolate_variables(value, variables)

        comparator = parse_comparator(operator, value)
        if comparator is None:
            continue

        block[name.strip().lower()] = comparator

    return block

def parse_comparator(operator : str, value : Any) -> Comparator:
    """Build a comparator from an operator string name and a value
    """

    op = Operators.get(operator.strip().lower(), None)
    assert op is not None, "Operator `%s` not recognised" % operator
    
    return Comparator(op, value)

def extract_directory(block : Dict[str, Comparator]) -> str:
    """Extract directory path from block
    """

    comp = block.get(GlobalKeys.DIRECTORY, None)
    if comp is None:
        return None
        
    assert isinstance(comp.value, (str, bytes,)) and comp.operator == Operator.EQUAL, \
        "Directory block must use operator `is` and a string value"
    
    # When on Windows, do like the Windowsians
    return comp.value.replace('/', os.path.sep)

def extract_filename(block : Dict[str, Comparator], current_directory : str) -> Tuple[str, str]:
    """Extract filename from block. May involve matching filesystem filenames.
    Returns a tuple of validated file path and filename match. May raise an
    AssertionError if file or directory not found. Files are compared case-insensitively.
    """

    comp = block.get(GlobalKeys.FILE, None)
    if comp is None:
        return (None, None,)
    
    assert isinstance(comp.value, (str, bytes,)) and comp.operator in (Operator.EQUAL, Operator.REGEX,), \
        "File block must use operator `is` or `matches` and a string value"
    
    assert os.path.isdir(current_directory), "Directory `%s` not found" % current_directory

    with_dir = lambda f: os.path.join(current_directory, f)
    is_file = lambda f: os.path.isfile(with_dir(f))

    time_sort = lambda f: os.path.getmtime(with_dir(f))
    files = sorted(filter(is_file, os.listdir(current_directory)), key=time_sort)

    filename = None
    match = None

    for f in reversed(files):
        if comp.operator == Operator.EQUAL:
            if comp.value.lower() == f.lower():
                match = filename = f
                break
        elif comp.operator == Operator.REGEX:
            match = comp.match(f)
            if match is not None:
                filename = f
                break

    assert filename is not None, "No matching file found for `%s` in `%s`" % (comp.value, current_directory,)
    assert is_file(filename), "File `%s` not found" % comp.value
    
    return (with_dir(filename), match,)

def cast_col(col : Any) -> int:
    """Column name or number to number
    """
    if isinstance(col, (str, bytes,)):
        try:
            col = column_index_from_string(col)
        except ValueError:
            pass
    assert col is None or isinstance(col, int), "%s is not a valid column name" % col
    return col

def contains_cell_match(block : Dict[str, Comparator], prefix : Callable = Prefix.none) -> bool:
    """Check if the block contains a cell reference (with optional prefix)
    """
    return prefix(CellMatchKeys.CELL_REFERENCE) in block or prefix(CellMatchKeys.VALUE) in block

def build_cell_match(block : Dict[str, Comparator], name : str = None, sheet : Comparator = None, prefix : Callable = Prefix.none) -> CellMatch:

    name = name if name is not None else block[MatchKeys.NAME].value if MatchKeys.NAME in block else None
    sheet = block[prefix(CellMatchKeys.SHEET)] if prefix(CellMatchKeys.SHEET) in block else sheet
    reference = block[prefix(CellMatchKeys.CELL_REFERENCE)].value if prefix(CellMatchKeys.CELL_REFERENCE) in block else None
    value = block[prefix(CellMatchKeys.VALUE)] if prefix(CellMatchKeys.VALUE) in block else None
    row_offset = block[prefix(CellMatchKeys.ROW_OFFSET)].value if prefix(CellMatchKeys.ROW_OFFSET) in block else 0
    col_offset = block[prefix(CellMatchKeys.COL_OFFSET)].value if prefix(CellMatchKeys.COL_OFFSET) in block else 0
    min_row = block[prefix(CellMatchKeys.MIN_ROW)].value if prefix(CellMatchKeys.MIN_ROW) in block else None
    max_row = block[prefix(CellMatchKeys.MAX_ROW)].value if prefix(CellMatchKeys.MAX_ROW) in block else None
    min_col = cast_col(block[prefix(CellMatchKeys.MIN_COL)].value) if prefix(CellMatchKeys.MIN_COL) in block else None
    max_col = cast_col(block[prefix(CellMatchKeys.MAX_COL)].value) if prefix(CellMatchKeys.MAX_COL) in block else None
    
    assert isinstance(name, (str, bytes,)), "Block name is required and must be a string"
    assert reference is None or isinstance(reference, (str, bytes,)), "Cell reference must be a string"
    assert row_offset is None or isinstance(row_offset, int), "Row offset must be a number"
    assert col_offset is None or isinstance(col_offset, int), "Column offset must be a number"
    assert min_row is None or isinstance(min_row, int), "Min row must be a number"
    assert max_row is None or isinstance(max_row, int), "Max row must be a number or column letter"
    assert min_col is None or isinstance(min_col, int), "Min column must be a number"
    assert max_col is None or isinstance(max_col, int), "Max column must be a number  or column letter"
    
    return CellMatch(
        name=name,
        sheet=sheet,
        reference=reference,
        value=value,
        row_offset=row_offset,
        col_offset=col_offset,
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    )

def build_range_match(block : Dict[str, Comparator]) -> CellMatch:

    name = block[MatchKeys.NAME].value if MatchKeys.NAME in block else None
    sheet = block[RangeMatchKeys.SHEET] if RangeMatchKeys.SHEET in block else None
    reference = block[RangeMatchKeys.TABLE_REFERENCE].value if RangeMatchKeys.TABLE_REFERENCE in block else None
    rows = block[RangeMatchKeys.ROWS].value if RangeMatchKeys.ROWS in block else None
    cols = block[RangeMatchKeys.COLS].value if RangeMatchKeys.COLS in block else None
    start_cell = build_cell_match(block, "%s:start" % name, sheet, Prefix.start) if contains_cell_match(block, Prefix.start) else None
    end_cell = build_cell_match(block, "%s:end" % name, sheet, Prefix.end) if contains_cell_match(block, Prefix.end) else None

    assert isinstance(name, (str, bytes,)), "Block name is required and must be a string"
    assert reference is None or isinstance(reference, (str, bytes,)), "Table reference must be a string"
    assert rows is None or isinstance(rows, int), "Rows must be a number"
    assert cols is None or isinstance(cols, int), "Columns must be a number"

    return RangeMatch(
        name=name,
        sheet=sheet,
        reference=reference,
        start_cell=start_cell,
        end_cell=end_cell,
        rows=rows,
        cols=cols,
    )

def extract_source_match(block : Dict[str, Comparator]) -> Match:
    """Create a CellMatch or RangeMatch from the block if possible
    """

    is_cell_match = MatchKeys.NAME in block and contains_cell_match(block)
    is_range_match = MatchKeys.NAME in block and (RangeMatchKeys.TABLE_REFERENCE in block or contains_cell_match(block, Prefix.start))

    assert not (is_cell_match and is_range_match), "Block refers to both a cell and a range"

    if is_cell_match:
        return build_cell_match(block)
    elif is_range_match:
        return build_range_match(block)
    
    return None

def extract_target(block : Dict[str, Comparator], source_match : Match) -> Target:
    """Create a Target from the block if possible
    """

    if not (TargetKeys.TARGET_CELL in block or TargetKeys.TARGET_TABLE in block):
        return None
    
    target_cell = block[TargetKeys.TARGET_CELL].value if TargetKeys.TARGET_CELL in block else None
    target_table = block[TargetKeys.TARGET_TABLE].value if TargetKeys.TARGET_TABLE in block else None

    align = block[TargetKeys.ALIGN].value if TargetKeys.ALIGN in block else False
    expand = block[TargetKeys.EXPAND].value if TargetKeys.EXPAND in block else False

    source_row = build_cell_match(block, "%s:source_row" % source_match.name, source_match.sheet, Prefix.source_row) if contains_cell_match(block, Prefix.source_row) else None
    source_col = build_cell_match(block, "%s:source_col" % source_match.name, source_match.sheet, Prefix.source_col) if contains_cell_match(block, Prefix.source_col) else None

    target_row = build_cell_match(block, "%s:target_row" % source_match.name, None, Prefix.target_row) if contains_cell_match(block, Prefix.target_row) else None
    target_col = build_cell_match(block, "%s:target_col" % source_match.name, None, Prefix.target_col) if contains_cell_match(block, Prefix.target_col) else None

    assert not (target_cell is not None and target_table is not None), "Only one of 'Target cell' and 'Target table' should be given"    
    assert target_cell is None or isinstance(target_cell, (str, bytes,)), "Target cell reference must be a string"
    assert target_table is None or isinstance(target_table, (str, bytes,)), "Target table reference must be a string"
    assert align is None or isinstance(align, bool), "Align must be a boolean (true/false)"
    assert expand is None or isinstance(align, bool), "Expand must be a boolean (true/false)"

    target_match = CellMatch("%s:target" % source_match.name, reference=target_cell) if target_cell is not None \
            else RangeMatch("%s:target" % source_match.name, reference=target_table)

    return Target(
        source=source_match,
        target=target_match,
        source_row=source_row,
        source_col=source_col,
        target_row=target_row,
        target_col=target_col,
        align=align,
        expand=expand,
    )

def interpolate_variables(value : str, variables : Dict[str, Any]) -> str:
    """Interpolate variables into the value string
    """

    if value is None or not isinstance(value, (str, bytes,)) or value == "":
        return value

    template = VariableTemplate(value)
    return template.safe_substitute(variables)
